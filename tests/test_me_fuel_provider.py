"""Tests for MeFuelProvider (data.gov.me — Montenegro government fuel prices)."""

from __future__ import annotations

import io
from unittest.mock import AsyncMock, MagicMock

import pytest
from aiohttp import ClientError

from custom_components.fuelcompare_ie.providers.base import ProviderError
from custom_components.fuelcompare_ie.providers.me_fuel import (
    MeFuelProvider,
    _CKAN_SEARCH_URL,
    _COL_EURODIESEL,
    _COL_EUROSUPER_95,
    _COL_EUROSUPER_98,
    _COL_LOZ_ULJE,
    _COL_TO_KEY,
    _HEADERS,
    _PRICE_ROW,
    _STATION_ID_ME,
    _parse_price,
    _parse_xlsx,
)


# ---------------------------------------------------------------------------
# CKAN API response fixtures
# ---------------------------------------------------------------------------

_XLSX_URL = "https://data.gov.me/dataset/gorivo/download/gorivo-2026-05.xlsx"
_MODIFIED = "2026-05-20T10:00:00"

_CKAN_SUCCESS_PAYLOAD: dict = {
    "success": True,
    "result": {
        "count": 1,
        "results": [
            {
                "id": "abc123",
                "metadata_modified": _MODIFIED,
                "resources": [
                    {
                        "format": "XLSX",
                        "url": _XLSX_URL,
                    }
                ],
            }
        ],
    },
}

_CKAN_NO_XLSX_PAYLOAD: dict = {
    "success": True,
    "result": {
        "count": 1,
        "results": [
            {
                "id": "abc123",
                "metadata_modified": _MODIFIED,
                "resources": [
                    {
                        "format": "PDF",
                        "url": "https://data.gov.me/dataset/gorivo/download/gorivo.pdf",
                    }
                ],
            }
        ],
    },
}

_CKAN_EMPTY_RESULTS_PAYLOAD: dict = {
    "success": True,
    "result": {"count": 0, "results": []},
}

_CKAN_FAILURE_PAYLOAD: dict = {
    "success": False,
    "result": {},
}

# Prices used in the synthetic XLSX fixture (matches confirmed May 2026 values)
_PRICE_E95 = 1.65
_PRICE_E98 = 1.68
_PRICE_EDIESEL = 1.69
_PRICE_LOZ = 1.73


# ---------------------------------------------------------------------------
# Synthetic XLSX builder
# ---------------------------------------------------------------------------


def _make_xlsx_bytes(
    e95: float | None = _PRICE_E95,
    e98: float | None = _PRICE_E98,
    diesel: float | None = _PRICE_EDIESEL,
    loz: float | None = _PRICE_LOZ,
    num_rows: int = 52,
) -> bytes:
    """Build a synthetic XLSX workbook with the expected layout.

    Creates a workbook whose active sheet has ``num_rows`` rows.
    When ``num_rows >= _PRICE_ROW`` (28), row 28 gets the supplied fuel
    price values in columns D-G.  When ``num_rows < _PRICE_ROW`` the MP
    row is intentionally omitted so the short-sheet path can be tested.
    """
    import openpyxl  # type: ignore[import]

    wb = openpyxl.Workbook()
    ws = wb.active

    # Anchor each row up to num_rows so openpyxl's max_row reflects
    # the intended sheet size.
    for r in range(1, num_rows + 1):
        ws.cell(row=r, column=1, value=f"row{r}")

    # Only write the MP row when the sheet is supposed to contain it.
    if num_rows >= _PRICE_ROW:
        ws.cell(row=_PRICE_ROW, column=1, value="MP")
        if e95 is not None:
            ws.cell(row=_PRICE_ROW, column=_COL_EUROSUPER_95, value=e95)
        if e98 is not None:
            ws.cell(row=_PRICE_ROW, column=_COL_EUROSUPER_98, value=e98)
        if diesel is not None:
            ws.cell(row=_PRICE_ROW, column=_COL_EURODIESEL, value=diesel)
        if loz is not None:
            ws.cell(row=_PRICE_ROW, column=_COL_LOZ_ULJE, value=loz)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Mock helpers
# ---------------------------------------------------------------------------


def _make_mock_response(
    status: int,
    json_data: dict | None = None,
    read_data: bytes | None = None,
    raise_on_raise_for_status: Exception | None = None,
) -> AsyncMock:
    """Build a mock aiohttp response usable as an async context manager."""
    mock_resp = AsyncMock()
    mock_resp.status = status
    mock_resp.json = AsyncMock(return_value=json_data if json_data is not None else {})
    mock_resp.read = AsyncMock(return_value=read_data if read_data is not None else b"")
    if raise_on_raise_for_status is not None:
        mock_resp.raise_for_status = MagicMock(side_effect=raise_on_raise_for_status)
    else:
        mock_resp.raise_for_status = MagicMock()
    mock_resp.__aenter__ = AsyncMock(return_value=mock_resp)
    mock_resp.__aexit__ = AsyncMock(return_value=False)
    return mock_resp


def _make_session(*responses: AsyncMock) -> MagicMock:
    """Return a mock session whose .get() call cycles through *responses*."""
    session = MagicMock()
    call_iter = iter(responses)

    def _get(*_args, **_kwargs):
        return next(call_iter)

    session.get = MagicMock(side_effect=_get)
    return session


def _default_provider() -> MeFuelProvider:
    """Return a MeFuelProvider with default settings."""
    return MeFuelProvider(station_id=_STATION_ID_ME)


# ---------------------------------------------------------------------------
# Provider metadata tests
# ---------------------------------------------------------------------------


def test_provider_country_is_me() -> None:
    """MeFuelProvider.COUNTRY is 'ME'."""
    assert MeFuelProvider.COUNTRY == "ME"


def test_provider_key_is_me_fuel() -> None:
    """MeFuelProvider.PROVIDER_KEY is 'me_fuel'."""
    assert MeFuelProvider.PROVIDER_KEY == "me_fuel"


def test_provider_label_contains_montenegro() -> None:
    """MeFuelProvider.LABEL contains 'Montenegro'."""
    assert "Montenegro" in MeFuelProvider.LABEL


def test_provider_config_mode_is_location() -> None:
    """CONFIG_MODE is 'location' (national average, no station selection)."""
    assert MeFuelProvider.CONFIG_MODE == "location"


def test_provider_station_lookup_mode_is_location_search() -> None:
    """STATION_LOOKUP_MODE is 'location_search'."""
    assert MeFuelProvider.STATION_LOOKUP_MODE == "location_search"


def test_provider_poll_interval_is_at_least_3600() -> None:
    """POLL_INTERVAL_SECONDS >= 3600 (data updates weekly/bi-weekly)."""
    assert MeFuelProvider.POLL_INTERVAL_SECONDS >= 3600


# ---------------------------------------------------------------------------
# Provider capabilities tests
# ---------------------------------------------------------------------------


def test_capabilities_include_diesel() -> None:
    assert "diesel" in MeFuelProvider.CAPABILITIES


def test_capabilities_include_unleaded() -> None:
    assert "unleaded" in MeFuelProvider.CAPABILITIES


def test_capabilities_include_premium_unleaded() -> None:
    assert "premium_unleaded" in MeFuelProvider.CAPABILITIES


def test_capabilities_include_kerosene() -> None:
    assert "kerosene" in MeFuelProvider.CAPABILITIES


def test_capabilities_include_lastupdated() -> None:
    assert "lastupdated" in MeFuelProvider.CAPABILITIES


def test_capabilities_include_coordinator_sentinels() -> None:
    caps = MeFuelProvider.CAPABILITIES
    assert "last_successful_fetch" in caps
    assert "data_fetch_problem" in caps


# ---------------------------------------------------------------------------
# Constructor tests
# ---------------------------------------------------------------------------


def test_constructor_default_station_id() -> None:
    """Default station_id is the country code 'ME'."""
    p = MeFuelProvider()
    assert p._station_id == "ME"


def test_constructor_accepts_coordinates() -> None:
    """Constructor stores optional lat/lng/radius for interface compat."""
    p = MeFuelProvider(station_id="ME", latitude=42.5, longitude=19.3, radius_km=50.0)
    assert p._latitude == pytest.approx(42.5)
    assert p._longitude == pytest.approx(19.3)
    assert p._radius_km == pytest.approx(50.0)


def test_constructor_extra_kwargs_do_not_raise() -> None:
    """Constructor absorbs unknown kwargs without raising."""
    MeFuelProvider(station_id="ME", county="Podgorica", unknown_param="x")


# ---------------------------------------------------------------------------
# Module constants tests
# ---------------------------------------------------------------------------


def test_ckan_search_url_points_to_data_gov_me() -> None:
    assert "data.gov.me" in _CKAN_SEARCH_URL
    assert "gorivo" in _CKAN_SEARCH_URL
    assert _CKAN_SEARCH_URL.startswith("https://")


def test_price_row_is_28() -> None:
    """_PRICE_ROW is 28 as documented in the XLSX layout."""
    assert _PRICE_ROW == 28


def test_col_to_key_maps_all_four_fuel_types() -> None:
    """_COL_TO_KEY covers all four fuel-type columns D-G."""
    assert _COL_TO_KEY[_COL_EUROSUPER_95] == "unleaded"
    assert _COL_TO_KEY[_COL_EUROSUPER_98] == "premium_unleaded"
    assert _COL_TO_KEY[_COL_EURODIESEL] == "diesel"
    assert _COL_TO_KEY[_COL_LOZ_ULJE] == "kerosene"


def test_station_id_me_constant() -> None:
    """_STATION_ID_ME is the string 'ME'."""
    assert _STATION_ID_ME == "ME"


def test_headers_include_user_agent() -> None:
    assert "User-Agent" in _HEADERS and _HEADERS["User-Agent"]


# ---------------------------------------------------------------------------
# _parse_price unit tests
# ---------------------------------------------------------------------------


def test_parse_price_normal_eur_per_litre() -> None:
    """_parse_price returns rounded float for normal value."""
    assert _parse_price(1.65) == pytest.approx(1.65)


def test_parse_price_rounds_to_3dp() -> None:
    assert _parse_price(1.65499) == pytest.approx(1.655)


def test_parse_price_none_returns_none() -> None:
    assert _parse_price(None) is None


def test_parse_price_zero_returns_none() -> None:
    assert _parse_price(0) is None


def test_parse_price_negative_returns_none() -> None:
    assert _parse_price(-1.5) is None


def test_parse_price_string_numeric() -> None:
    assert _parse_price("1.65") == pytest.approx(1.65)


def test_parse_price_non_numeric_string_returns_none() -> None:
    assert _parse_price("N/A") is None


def test_parse_price_cents_guard_above_10() -> None:
    """Values >10 are divided by 100 (guard for accidental cents)."""
    assert _parse_price(165.0) == pytest.approx(1.65)


def test_parse_price_integer_input() -> None:
    assert _parse_price(2) == pytest.approx(2.0)


# ---------------------------------------------------------------------------
# _parse_xlsx unit tests
# ---------------------------------------------------------------------------


def test_parse_xlsx_success_all_four_fuel_types() -> None:
    """_parse_xlsx extracts all four fuel type prices from row 28."""
    data = _parse_xlsx(_make_xlsx_bytes())
    assert data["unleaded"] == pytest.approx(_PRICE_E95)
    assert data["premium_unleaded"] == pytest.approx(_PRICE_E98)
    assert data["diesel"] == pytest.approx(_PRICE_EDIESEL)
    assert data["kerosene"] == pytest.approx(_PRICE_LOZ)


def test_parse_xlsx_returns_none_for_missing_cells() -> None:
    """_parse_xlsx returns None for fuel types whose cells are blank."""
    data = _parse_xlsx(_make_xlsx_bytes(e95=None, e98=None))
    assert data["unleaded"] is None
    assert data["premium_unleaded"] is None
    assert data["diesel"] == pytest.approx(_PRICE_EDIESEL)
    assert data["kerosene"] == pytest.approx(_PRICE_LOZ)


def test_parse_xlsx_raises_provider_error_for_invalid_bytes() -> None:
    """_parse_xlsx raises ProviderError when content is not a valid XLSX."""
    with pytest.raises(ProviderError, match="MeFuel"):
        _parse_xlsx(b"this is not an xlsx file at all")


def test_parse_xlsx_short_sheet_returns_none_prices() -> None:
    """_parse_xlsx returns all-None prices when the sheet has < 28 rows."""
    data = _parse_xlsx(_make_xlsx_bytes(num_rows=10))
    assert data["unleaded"] is None
    assert data["premium_unleaded"] is None
    assert data["diesel"] is None
    assert data["kerosene"] is None


def test_parse_xlsx_all_capability_keys_present() -> None:
    """_parse_xlsx result contains exactly the four expected keys."""
    data = _parse_xlsx(_make_xlsx_bytes())
    assert set(data.keys()) == {"unleaded", "premium_unleaded", "diesel", "kerosene"}


# ---------------------------------------------------------------------------
# async_fetch — success path
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_async_fetch_success_returns_all_fuel_prices() -> None:
    """async_fetch returns a populated StationData with all four prices."""
    xlsx_bytes = _make_xlsx_bytes()
    ckan_resp = _make_mock_response(200, json_data=_CKAN_SUCCESS_PAYLOAD)
    xlsx_resp = _make_mock_response(200, read_data=xlsx_bytes)
    session = _make_session(ckan_resp, xlsx_resp)

    provider = _default_provider()
    data = await provider.async_fetch(session, _STATION_ID_ME)

    assert data["unleaded"] == pytest.approx(_PRICE_E95)
    assert data["premium_unleaded"] == pytest.approx(_PRICE_E98)
    assert data["diesel"] == pytest.approx(_PRICE_EDIESEL)
    assert data["kerosene"] == pytest.approx(_PRICE_LOZ)


@pytest.mark.asyncio
async def test_async_fetch_success_populates_lastupdated() -> None:
    """async_fetch returns lastupdated from the CKAN metadata_modified field."""
    xlsx_bytes = _make_xlsx_bytes()
    ckan_resp = _make_mock_response(200, json_data=_CKAN_SUCCESS_PAYLOAD)
    xlsx_resp = _make_mock_response(200, read_data=xlsx_bytes)
    session = _make_session(ckan_resp, xlsx_resp)

    provider = _default_provider()
    data = await provider.async_fetch(session, _STATION_ID_ME)

    assert data["lastupdated"] == _MODIFIED


@pytest.mark.asyncio
async def test_async_fetch_source_station_id_is_me() -> None:
    """async_fetch sets source_station_id to 'ME'."""
    xlsx_bytes = _make_xlsx_bytes()
    ckan_resp = _make_mock_response(200, json_data=_CKAN_SUCCESS_PAYLOAD)
    xlsx_resp = _make_mock_response(200, read_data=xlsx_bytes)
    session = _make_session(ckan_resp, xlsx_resp)

    provider = _default_provider()
    data = await provider.async_fetch(session, _STATION_ID_ME)

    assert data["source_station_id"] == "ME"


@pytest.mark.asyncio
async def test_async_fetch_makes_exactly_two_requests() -> None:
    """async_fetch makes exactly 2 HTTP requests: CKAN search + XLSX download."""
    xlsx_bytes = _make_xlsx_bytes()
    ckan_resp = _make_mock_response(200, json_data=_CKAN_SUCCESS_PAYLOAD)
    xlsx_resp = _make_mock_response(200, read_data=xlsx_bytes)
    session = _make_session(ckan_resp, xlsx_resp)

    provider = _default_provider()
    await provider.async_fetch(session, _STATION_ID_ME)

    assert session.get.call_count == 2


# ---------------------------------------------------------------------------
# async_fetch — CKAN error handling
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_async_fetch_raises_on_ckan_http_error() -> None:
    """async_fetch raises ProviderError when the CKAN API call fails."""
    ckan_resp = _make_mock_response(
        503,
        raise_on_raise_for_status=ClientError("service unavailable"),
    )
    session = _make_session(ckan_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError, match="MeFuel"):
        await provider.async_fetch(session, _STATION_ID_ME)


@pytest.mark.asyncio
async def test_async_fetch_raises_on_ckan_network_error() -> None:
    """async_fetch raises ProviderError on network failure."""
    session = MagicMock()
    session.get = MagicMock(side_effect=ClientError("connection refused"))

    provider = _default_provider()
    with pytest.raises(ProviderError):
        await provider.async_fetch(session, _STATION_ID_ME)


@pytest.mark.asyncio
async def test_async_fetch_raises_when_ckan_returns_success_false() -> None:
    """async_fetch raises ProviderError when CKAN returns success=false."""
    ckan_resp = _make_mock_response(200, json_data=_CKAN_FAILURE_PAYLOAD)
    session = _make_session(ckan_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError, match="success=false"):
        await provider.async_fetch(session, _STATION_ID_ME)


@pytest.mark.asyncio
async def test_async_fetch_raises_when_ckan_results_empty() -> None:
    """async_fetch raises ProviderError when CKAN returns no datasets."""
    ckan_resp = _make_mock_response(200, json_data=_CKAN_EMPTY_RESULTS_PAYLOAD)
    session = _make_session(ckan_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError, match="no datasets"):
        await provider.async_fetch(session, _STATION_ID_ME)


@pytest.mark.asyncio
async def test_async_fetch_raises_when_no_xlsx_resource_found() -> None:
    """async_fetch raises ProviderError when dataset has no XLSX resource."""
    ckan_resp = _make_mock_response(200, json_data=_CKAN_NO_XLSX_PAYLOAD)
    session = _make_session(ckan_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError, match="No XLSX"):
        await provider.async_fetch(session, _STATION_ID_ME)


# ---------------------------------------------------------------------------
# async_fetch — XLSX download error handling
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_async_fetch_raises_on_xlsx_download_http_error() -> None:
    """async_fetch raises ProviderError when XLSX download returns HTTP error."""
    ckan_resp = _make_mock_response(200, json_data=_CKAN_SUCCESS_PAYLOAD)
    xlsx_resp = _make_mock_response(
        404,
        raise_on_raise_for_status=ClientError("not found"),
    )
    session = _make_session(ckan_resp, xlsx_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError, match="MeFuel"):
        await provider.async_fetch(session, _STATION_ID_ME)


@pytest.mark.asyncio
async def test_async_fetch_raises_on_invalid_xlsx_content() -> None:
    """async_fetch raises ProviderError when the downloaded file is not XLSX."""
    ckan_resp = _make_mock_response(200, json_data=_CKAN_SUCCESS_PAYLOAD)
    xlsx_resp = _make_mock_response(200, read_data=b"<html>not an xlsx</html>")
    session = _make_session(ckan_resp, xlsx_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError):
        await provider.async_fetch(session, _STATION_ID_ME)


# ---------------------------------------------------------------------------
# async_fetch — XLSX URL fallback (url ends in .xlsx without format=XLSX)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_async_fetch_falls_back_to_url_extension_when_format_missing() -> None:
    """async_fetch finds XLSX by .xlsx URL extension when format field is blank."""
    payload_no_format: dict = {
        "success": True,
        "result": {
            "count": 1,
            "results": [
                {
                    "id": "abc456",
                    "metadata_modified": _MODIFIED,
                    "resources": [
                        {
                            "format": "",  # format field absent / blank
                            "url": _XLSX_URL,  # URL ends in .xlsx
                        }
                    ],
                }
            ],
        },
    }
    xlsx_bytes = _make_xlsx_bytes()
    ckan_resp = _make_mock_response(200, json_data=payload_no_format)
    xlsx_resp = _make_mock_response(200, read_data=xlsx_bytes)
    session = _make_session(ckan_resp, xlsx_resp)

    provider = _default_provider()
    data = await provider.async_fetch(session, _STATION_ID_ME)

    assert data["diesel"] == pytest.approx(_PRICE_EDIESEL)


# ---------------------------------------------------------------------------
# async_fetch_station_name
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_async_fetch_station_name_always_returns_none() -> None:
    """async_fetch_station_name always returns None (location-mode provider)."""
    session = MagicMock()
    provider = _default_provider()
    name = await provider.async_fetch_station_name(session, _STATION_ID_ME)
    assert name is None


@pytest.mark.asyncio
async def test_async_fetch_station_name_makes_no_requests() -> None:
    """async_fetch_station_name does not call session.get."""
    session = MagicMock()
    provider = _default_provider()
    await provider.async_fetch_station_name(session, _STATION_ID_ME)
    session.get.assert_not_called()


# ---------------------------------------------------------------------------
# async_list_stations
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_async_list_stations_returns_single_entry() -> None:
    """async_list_stations returns exactly one entry for the national average."""
    session = MagicMock()
    provider = _default_provider()
    result = await provider.async_list_stations(session)

    assert len(result) == 1


@pytest.mark.asyncio
async def test_async_list_stations_station_id_is_me() -> None:
    """async_list_stations returns station_id 'ME'."""
    session = MagicMock()
    provider = _default_provider()
    result = await provider.async_list_stations(session)

    station_id, _label = result[0]
    assert station_id == "ME"


@pytest.mark.asyncio
async def test_async_list_stations_label_contains_montenegro() -> None:
    """async_list_stations label mentions Montenegro."""
    session = MagicMock()
    provider = _default_provider()
    result = await provider.async_list_stations(session)

    _sid, label = result[0]
    assert "Montenegro" in label


@pytest.mark.asyncio
async def test_async_list_stations_makes_no_requests() -> None:
    """async_list_stations does not make any HTTP requests."""
    session = MagicMock()
    provider = _default_provider()
    await provider.async_list_stations(session, lat=42.5, lng=19.3, radius_km=50.0)
    session.get.assert_not_called()


# ---------------------------------------------------------------------------
# _fetch_xlsx_url — unit tests via internal helper
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_fetch_xlsx_url_returns_url_and_modified() -> None:
    """_fetch_xlsx_url returns (xlsx_url, metadata_modified) on success."""
    ckan_resp = _make_mock_response(200, json_data=_CKAN_SUCCESS_PAYLOAD)
    session = _make_session(ckan_resp)

    provider = _default_provider()
    url, modified = await provider._fetch_xlsx_url(session)

    assert url == _XLSX_URL
    assert modified == _MODIFIED


@pytest.mark.asyncio
async def test_fetch_xlsx_url_raises_on_http_error() -> None:
    """_fetch_xlsx_url raises ProviderError on HTTP failure."""
    ckan_resp = _make_mock_response(
        500,
        raise_on_raise_for_status=ClientError("server error"),
    )
    session = _make_session(ckan_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError):
        await provider._fetch_xlsx_url(session)


@pytest.mark.asyncio
async def test_fetch_xlsx_url_raises_when_no_results() -> None:
    """_fetch_xlsx_url raises ProviderError when CKAN returns no results."""
    ckan_resp = _make_mock_response(200, json_data=_CKAN_EMPTY_RESULTS_PAYLOAD)
    session = _make_session(ckan_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError):
        await provider._fetch_xlsx_url(session)


@pytest.mark.asyncio
async def test_fetch_xlsx_url_raises_when_no_xlsx_resource() -> None:
    """_fetch_xlsx_url raises ProviderError when no XLSX resource URL is found."""
    ckan_resp = _make_mock_response(200, json_data=_CKAN_NO_XLSX_PAYLOAD)
    session = _make_session(ckan_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError, match="No XLSX"):
        await provider._fetch_xlsx_url(session)


# ---------------------------------------------------------------------------
# _download_xlsx — unit tests via internal helper
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_download_xlsx_returns_bytes() -> None:
    """_download_xlsx returns the raw bytes from the HTTP response."""
    expected = b"PK\x03\x04fake xlsx content"
    xlsx_resp = _make_mock_response(200, read_data=expected)
    session = _make_session(xlsx_resp)

    provider = _default_provider()
    result = await provider._download_xlsx(session, _XLSX_URL)

    assert result == expected


@pytest.mark.asyncio
async def test_download_xlsx_raises_on_http_error() -> None:
    """_download_xlsx raises ProviderError on HTTP failure."""
    xlsx_resp = _make_mock_response(
        403,
        raise_on_raise_for_status=ClientError("forbidden"),
    )
    session = _make_session(xlsx_resp)

    provider = _default_provider()
    with pytest.raises(ProviderError, match="MeFuel"):
        await provider._download_xlsx(session, _XLSX_URL)


@pytest.mark.asyncio
async def test_download_xlsx_raises_on_network_error() -> None:
    """_download_xlsx raises ProviderError on network failure."""
    session = MagicMock()
    session.get = MagicMock(side_effect=ClientError("timeout"))

    provider = _default_provider()
    with pytest.raises(ProviderError):
        await provider._download_xlsx(session, _XLSX_URL)


# ---------------------------------------------------------------------------
# Integration: provider registered in PROVIDER_REGISTRY
# ---------------------------------------------------------------------------


def test_provider_registered_in_registry() -> None:
    """MeFuelProvider is registered in the provider registry."""
    from custom_components.fuelcompare_ie.providers import PROVIDER_REGISTRY

    assert "me_fuel" in PROVIDER_REGISTRY
    assert PROVIDER_REGISTRY["me_fuel"] is MeFuelProvider


def test_provider_registry_get_provider_class() -> None:
    """get_provider_class returns MeFuelProvider for key 'me_fuel'."""
    from custom_components.fuelcompare_ie.providers import get_provider_class

    cls = get_provider_class("me_fuel")
    assert cls is MeFuelProvider
