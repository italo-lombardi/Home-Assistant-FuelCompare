"""MeFuelProvider — Montenegro government fuel prices via data.gov.me (CKAN).

Source: Montenegro Ministry of Energy and Mining (Ministarstvo energetike
i rudarstva).  Fuel prices are government-mandated maximum retail prices
published as XLSX workbooks on the national open-data portal data.gov.me,
which is a CKAN instance.

Two-step fetch process
----------------------
1. CKAN package_search API — returns the most recently modified dataset
   matching the query ``gorivo`` (fuel).  The response includes a
   ``resources`` list; the first resource whose ``format`` is ``XLSX``
   (case-insensitive) contains the ``url`` of the downloadable workbook.

2. XLSX download & parse — the workbook is a government calculation sheet
   spanning columns A:W, rows 1:52.  Retail prices (Zaokružena maksimalna
   maloprodajna cijena — rounded maximum retail price) are in row 28,
   identified in column A by the label ``"MP"`` (Maloprodajna cijena).
   Columns D–G contain the four fuel types in this fixed order:

     Column D — EUROSUPER 95
     Column E — EUROSUPER 98
     Column F — EURODIESEL
     Column G — LOŽ ULJE (heating oil / lož ulje)

All four prices are in EUR/litre.  The government publishes updates
approximately weekly to bi-weekly; there is no fixed update day.

No station-level data is available from any free/open source for
Montenegro.  This provider therefore operates in STATION_LOOKUP_MODE='global_list'
with a single virtual station whose station_id is the country code ``'ME'``.

CKAN API endpoint
-----------------
GET https://data.gov.me/api/3/action/package_search
  ?q=gorivo&rows=1&sort=metadata_modified+desc

Returns JSON:
{
  "success": true,
  "result": {
    "results": [
      {
        "metadata_modified": "2026-05-...",
        "resources": [
          {
            "format": "XLSX",
            "url": "https://data.gov.me/dataset/.../download/gorivo-...xlsx"
          }
        ]
      }
    ]
  }
}

XLSX layout (columns are 0-indexed internally after openpyxl read):
  Workbook sheet 1 (active sheet), row 28 is the ``MP`` row.
  openpyxl is used to parse the binary XLSX content returned by the
  download URL.  The file is read from an in-memory BytesIO buffer so no
  temporary files are created.

StationData mapping
-------------------
XLSX column  Fuel type       → StationData key
-----------  ---------          ----------------
D (col 4)    EUROSUPER 95    → unleaded
E (col 5)    EUROSUPER 98    → premium_unleaded
F (col 6)    EURODIESEL      → diesel
G (col 7)    LOŽ ULJE        → kerosene  (heating oil)

All prices are stored as EUR/litre floats, rounded to 3 decimal places.

Error handling
--------------
- If the CKAN API call fails, ProviderError is raised.
- If no XLSX resource URL is found in the dataset, ProviderError is raised.
- If the XLSX download fails or cannot be parsed, ProviderError is raised.
- Individual cell parse failures return None for that fuel type (logged
  at DEBUG level) rather than aborting the whole fetch.
"""

from __future__ import annotations

import asyncio
import functools
import io
import logging
import re
from typing import Any, ClassVar

from aiohttp import ClientSession, ClientTimeout

from ..const import UA_HEADER, API_TIMEOUT
from .base import BaseProvider, ProviderError, StationData

_LOGGER = logging.getLogger(__name__)

# ── API / URL constants ───────────────────────────────────────────────────────

_CKAN_SEARCH_URL = (
    "https://data.gov.me/api/3/action/package_search"
    "?q=gorivo&rows=5&sort=metadata_modified+desc"
)

_HEADERS: dict[str, str] = {
    "User-Agent": UA_HEADER,
    "Accept": "application/json",
}

_TIMEOUT = ClientTimeout(total=max(API_TIMEOUT * 3, 30))  # XLSX download can be slow

# ── XLSX layout constants ─────────────────────────────────────────────────────

# Row number (1-indexed, as used by openpyxl) for the final retail price.
# Row 28 is labelled "MP" in column A of the government workbook.
_PRICE_ROW = 28

# Column indices (1-indexed, as used by openpyxl) for the four fuel types.
# D=4, E=5, F=6, G=7.
_COL_EUROSUPER_95 = 4  # D — EUROSUPER 95
_COL_EUROSUPER_98 = 5  # E — EUROSUPER 98
_COL_EURODIESEL = 6  # F — EURODIESEL
_COL_LOZ_ULJE = 7  # G — LOŽ ULJE (heating oil)

# Column-to-StationData-key mapping (column index → StationData key).
_COL_TO_KEY: dict[int, str] = {
    _COL_EUROSUPER_95: "unleaded",
    _COL_EUROSUPER_98: "premium_unleaded",
    _COL_EURODIESEL: "diesel",
    _COL_LOZ_ULJE: "kerosene",
}

# Virtual station_id used for the single national-average station.
_STATION_ID_ME = "ME"


class MeFuelProvider(BaseProvider):
    """Fetch Montenegrin government maximum retail fuel prices.

    The source is the Montenegro open-data portal (data.gov.me, CKAN).
    The Ministry of Energy and Mining publishes government-mandated maximum
    retail prices for EUROSUPER 95, EUROSUPER 98, EURODIESEL, and LOŽ ULJE
    (heating oil) as XLSX workbooks.

    There is no station-level data available.  STATION_LOOKUP_MODE is
    ``'global_list'`` with the virtual station ``'ME'`` so the coordinator
    tracks a single national-average entry per integration instance.

    The fetch process has two steps:
      1. Query the CKAN package_search endpoint to get the latest XLSX URL.
      2. Download and parse the XLSX to extract row 28 (MP prices) for
         columns D–G.

    Usage
    -----
    provider = MeFuelProvider(station_id="ME")
    data = await provider.async_fetch(session, "ME")
    """

    COUNTRY = "ME"
    PROVIDER_KEY = "me_fuel"
    LABEL = "Min. of Energy (Montenegro)"
    CONFIG_MODE = "station_id"
    STATION_LOOKUP_MODE = "global_list"
    POLL_INTERVAL_SECONDS = 43200
    STATION_PAGE_URL: ClassVar[str] = (
        "https://data.gov.me"  # 12 hours; source updates approx weekly/bi-weekly
    )

    CAPABILITIES: ClassVar[frozenset[str]] = frozenset(
        {
            # Fuel prices
            "unleaded",  # EUROSUPER 95
            "premium_unleaded",  # EUROSUPER 98
            "diesel",  # EURODIESEL
            "kerosene",  # LOŽ ULJE (heating oil)
            # Timing
            "lastupdated",
        }
    )

    def __init__(self, station_id: str = _STATION_ID_ME) -> None:
        """Initialise the provider.

        Args:
            station_id:  Always ``'ME'`` for this national-average provider.
        """
        self._station_id = station_id

    # ── Public interface ──────────────────────────────────────────────────────

    async def async_fetch(
        self,
        session: ClientSession,
        station_id: str,
    ) -> StationData:
        """Fetch the latest government fuel prices from data.gov.me.

        Performs a two-step fetch:
          1. CKAN package_search to get the latest XLSX download URL (scans
             up to 5 results).  Falls back to description-text parsing when
             no XLSX is found across all rows.
          2. XLSX download and parse to extract the MP (Maloprodajna cijena)
             row for EUROSUPER 95, EUROSUPER 98, EURODIESEL, and LOŽ ULJE.

        Args:
            session:    aiohttp ClientSession provided by the coordinator.
            station_id: Ignored (always the virtual station ``'ME'``).

        Returns:
            StationData dict with all CAPABILITIES keys populated (prices may
            be None if the XLSX cell is empty or non-numeric).

        Raises:
            ProviderError: CKAN API call failed, no XLSX resource found and
                           description parse also yielded no prices.
        """
        xlsx_url, modified, fallback_description = await self._fetch_xlsx_url(session)

        if xlsx_url is not None:
            xlsx_bytes = await self._download_xlsx(session, xlsx_url)
            prices = await _parse_xlsx(xlsx_bytes)
        else:
            _LOGGER.debug(
                "MeFuel: no XLSX found across all rows; falling back to description parse"
            )
            prices = _parse_prices_from_description(fallback_description or "")
            if not any(v is not None for v in prices.values()):
                raise ProviderError(
                    "MeFuel: no XLSX resource found and description parse yielded no prices. "
                    "The data.gov.me portal may have changed its format."
                )

        data: StationData = {
            "unleaded": prices.get("unleaded"),
            "premium_unleaded": prices.get("premium_unleaded"),
            "diesel": prices.get("diesel"),
            "kerosene": prices.get("kerosene"),
            "lastupdated": modified,
            "source_station_id": _STATION_ID_ME,
        }

        _LOGGER.debug(
            "MeFuel parsed prices: unleaded=%s premium_unleaded=%s "
            "diesel=%s kerosene=%s lastupdated=%s",
            data.get("unleaded"),
            data.get("premium_unleaded"),
            data.get("diesel"),
            data.get("kerosene"),
            modified,
        )

        return data

    async def async_fetch_station_name(
        self,
        session: ClientSession,
        station_id: str,
    ) -> str | None:
        """Return None — location-mode providers use auto-generated titles.

        The config flow uses the auto-generated ``'Montenegro (national avg)'``
        title for this provider.

        Args:
            session:    aiohttp ClientSession (not used).
            station_id: Ignored.
        """
        return None

    async def async_list_stations(
        self,
        session: ClientSession,
        **kwargs: Any,
    ) -> list[tuple[str, str]]:
        """Return a single virtual station entry for the location picker.

        Since Montenegro has no station-level data, a single entry
        representing the national average is returned.  The config flow
        will display this as the only option.

        Args:
            session:   aiohttp ClientSession.
            lat:       Ignored (no station-level data).
            lng:       Ignored (no station-level data).
            radius_km: Ignored (no station-level data).

        Returns:
            List with one entry: ``[("ME", "Montenegro — national average")]``.
        """
        return [(_STATION_ID_ME, "Montenegro — national average (max retail prices)")]

    # ── Internal helpers ──────────────────────────────────────────────────────

    async def _fetch_xlsx_url(
        self, session: ClientSession
    ) -> tuple[str | None, str | None, str | None]:
        """Query the CKAN package_search API and return the XLSX download URL.

        Scans up to 5 results (rows=5) to find the first dataset that has a
        valid XLSX resource.  Returns a 3-tuple:
          - xlsx_url:             URL string, or None if no XLSX found
          - metadata_modified:    ISO 8601 string from the XLSX dataset, or
                                  from results[0] when falling back
          - fallback_description: description text of results[0] for use when
                                  no XLSX was found; None otherwise

        Args:
            session: aiohttp ClientSession.

        Raises:
            ProviderError: CKAN API call failed or returned no results.
        """
        _LOGGER.debug("MeFuel: querying CKAN package_search for latest gorivo dataset")
        try:
            async with session.get(
                _CKAN_SEARCH_URL,
                headers=_HEADERS,
                timeout=_TIMEOUT,
            ) as response:
                response.raise_for_status()
                payload: dict[str, Any] = await response.json(content_type=None)
        except ProviderError:
            raise
        except Exception as err:
            raise ProviderError(
                f"MeFuel: CKAN package_search request failed: {err}"
            ) from err

        if not payload.get("success"):
            raise ProviderError(
                "MeFuel: CKAN package_search returned success=false. "
                f"Response: {payload}"
            )

        results: list[dict] = payload.get("result", {}).get("results") or []
        if not results:
            raise ProviderError(
                "MeFuel: CKAN package_search returned no datasets for query 'gorivo'. "
                "The data.gov.me portal may have restructured its dataset naming."
            )

        # Scan all returned results for the first valid XLSX resource.
        for dataset in results:
            modified: str | None = dataset.get("metadata_modified") or None
            resources: list[dict] = dataset.get("resources") or []
            xlsx_url: str | None = None

            for resource in resources:
                fmt: str = (resource.get("format") or "").upper()
                url: str = resource.get("url") or ""
                if fmt == "XLSX" and url:
                    xlsx_url = url
                    break

            if not xlsx_url:
                for resource in resources:
                    url = resource.get("url") or ""
                    if url.lower().endswith(".xlsx"):
                        xlsx_url = url
                        break

            if xlsx_url:
                if not xlsx_url.startswith("https://data.gov.me/"):
                    raise ProviderError(
                        f"Refusing to download from untrusted host: {xlsx_url}"
                    )
                _LOGGER.debug(
                    "MeFuel: found XLSX URL: %s (modified=%s)", xlsx_url, modified
                )
                return xlsx_url, modified, None

        # No XLSX found across all rows — return description of most recent entry.
        first_dataset = results[0]
        fallback_modified: str | None = first_dataset.get("metadata_modified") or None
        fallback_description: str | None = first_dataset.get("notes") or None
        _LOGGER.debug(
            "MeFuel: no XLSX found in %d results; will attempt description fallback",
            len(results),
        )
        return None, fallback_modified, fallback_description

    async def _download_xlsx(self, session: ClientSession, xlsx_url: str) -> bytes:
        """Download the XLSX file and return its raw bytes.

        Args:
            session:  aiohttp ClientSession.
            xlsx_url: Direct download URL for the XLSX workbook.

        Returns:
            Raw XLSX bytes.

        Raises:
            ProviderError: HTTP error or network failure during download.
        """
        _LOGGER.debug("MeFuel: downloading XLSX from %s", xlsx_url)
        try:
            async with session.get(
                xlsx_url,
                headers=_HEADERS,
                timeout=_TIMEOUT,
            ) as response:
                response.raise_for_status()
                return await response.read()
        except ProviderError:
            raise
        except Exception as err:
            raise ProviderError(
                f"MeFuel: failed to download XLSX from {xlsx_url!r}: {err}"
            ) from err


# ── Module-level helpers ──────────────────────────────────────────────────────


def _parse_price(raw: Any) -> float | None:
    """Parse a raw cell value from the government XLSX workbook.

    Prices are already in EUR/litre (e.g. 1.650 or 1.69).  Returns None for
    None, zero, or non-numeric values.

    Args:
        raw: Cell value from openpyxl (float, int, str, or None).

    Returns:
        Normalised price as EUR/litre (rounded to 3 dp), or None.
    """
    if raw is None:
        return None
    try:
        val = float(raw)
    except (ValueError, TypeError):
        return None
    if val <= 0:
        return None
    # Prices in the Montenegrin XLSX are already EUR/litre (~1.6–2.0).
    # Apply the standard >10 guard in case the XLSX ever switches to cents.
    if val > 10:
        val = val / 100.0
    return round(val, 3)


def _parse_prices_from_description(description: str) -> dict[str, float | None]:
    """Parse fuel prices from a CKAN dataset description text.

    Used as a fallback when no XLSX resource is found in any of the returned
    datasets.  The Montenegrin portal sometimes publishes a dataset entry that
    contains only descriptive text (no attached file) for new price updates.

    Matches patterns like ``EUROSUPER 95 1,66 eur`` where the price uses a
    comma as the decimal separator.

    Args:
        description: The ``notes`` field of the most recent CKAN dataset.

    Returns:
        Dict with keys ``'unleaded'``, ``'premium_unleaded'``, ``'diesel'``,
        ``'kerosene'``; each value is a float (EUR/litre) or None.
    """
    prices: dict[str, float | None] = {
        "unleaded": None,
        "premium_unleaded": None,
        "diesel": None,
        "kerosene": None,
    }

    _FUEL_PATTERNS: list[tuple[str, str]] = [
        (r"EUROSUPER\s+95", "unleaded"),
        (r"EUROSUPER\s+98", "premium_unleaded"),
        (r"EURODI(?:EZEL|ESEL|ZIEL)", "diesel"),
        (r"LO[ŽZ]\s+ULJE", "kerosene"),
    ]

    for fuel_pattern, key in _FUEL_PATTERNS:
        match = re.search(
            fuel_pattern + r"\s+(\d+[,\.]\d+)\s+eur",
            description,
            re.IGNORECASE,
        )
        if match:
            raw = match.group(1).replace(",", ".")
            try:
                val = float(raw)
            except ValueError:
                continue
            if val > 0:
                prices[key] = round(val, 3)

    _LOGGER.debug(
        "MeFuel: description-parsed prices: %s",
        {k: v for k, v in prices.items() if v is not None},
    )
    return prices


async def _parse_xlsx(xlsx_bytes: bytes) -> dict[str, float | None]:
    """Parse the government XLSX workbook and return fuel prices.

    Locates row 28 (the ``MP`` row — Zaokružena maksimalna maloprodajna
    cijena) and reads columns D–G for the four fuel types.

    The row is identified by its 1-based row number (28) as documented.
    If the sheet has fewer than 28 rows, all prices return as None rather
    than raising an exception.

    Args:
        xlsx_bytes: Raw XLSX file content as bytes.

    Returns:
        Dict with keys ``'unleaded'``, ``'premium_unleaded'``, ``'diesel'``,
        ``'kerosene'``; each value is a float (EUR/litre) or None.

    Raises:
        ProviderError: The bytes cannot be parsed as a valid XLSX workbook.
    """
    try:
        import openpyxl  # local import — openpyxl is not always installed  # noqa: PLC0415

        wb = await asyncio.get_running_loop().run_in_executor(
            None,
            functools.partial(
                openpyxl.load_workbook,
                io.BytesIO(xlsx_bytes),
                read_only=True,
                data_only=True,
            ),
        )
        ws = wb.active
    except Exception as err:
        raise ProviderError(f"MeFuel: failed to parse XLSX workbook: {err}") from err

    prices: dict[str, float | None] = {
        "unleaded": None,
        "premium_unleaded": None,
        "diesel": None,
        "kerosene": None,
    }

    try:
        # ws.max_row is None for empty sheets; guard against it.
        max_row = ws.max_row or 0  # type: ignore[union-attr]
        if max_row < _PRICE_ROW:
            _LOGGER.warning(
                "MeFuel: XLSX has only %d rows; expected at least %d for MP row.",
                max_row,
                _PRICE_ROW,
            )
            return prices

        # openpyxl cell() uses 1-based row and column indices.
        for col_idx, data_key in _COL_TO_KEY.items():
            cell = ws.cell(row=_PRICE_ROW, column=col_idx)  # type: ignore[union-attr]
            val = cell.value
            parsed = _parse_price(val)
            if parsed is None and val is not None:
                _LOGGER.debug(
                    "MeFuel: could not parse cell (%d, %d) value=%r for key=%s",
                    _PRICE_ROW,
                    col_idx,
                    val,
                    data_key,
                )
            prices[data_key] = parsed
    finally:
        try:
            wb.close()  # type: ignore[union-attr]
        except Exception:  # noqa: BLE001
            pass

    _LOGGER.debug(
        "MeFuel: parsed XLSX prices: %s",
        {k: v for k, v in prices.items() if v is not None},
    )
    return prices
