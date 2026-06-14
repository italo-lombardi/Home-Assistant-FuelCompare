"""EuOilBulletinProvider — EU weekly fuel prices from the EC Oil Bulletin.

Source: European Commission — Weekly Oil Bulletin
  https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en

Data format: Excel (.xlsx), downloaded via a stable UUID URL.
No API key required; plain HTTPS GET returns the current week's data.

UUID stability
--------------
The EC uses permanent UUIDs for the download links.  The URL filename may
show an older date but the file content always contains the latest weekly
submission.  This has been confirmed: UUID 264c2d0f-… with a filename of
2024-02-19 served data dated 2026-06-08.

Two files are available:
  prices_with_taxes    UUID: 264c2d0f-f161-4ea3-a777-78faae59bea0  (~14 KB)
  prices_without_taxes UUID: 78311f92-68f8-4b82-b5cf-1293beeaae77  (~14 KB)
  history (since 2005) UUID: 906e60ca-8b6a-44e7-8589-652854d2fd3f  (~4.4 MB)

This provider uses prices_with_taxes by default.

Excel structure (Sheet1)
------------------------
Row 1: header labels (language titles)
Row 2: column identifiers / units
Row 3+: one country per row

Column layout (0-indexed after openpyxl 1-indexed read, rows start at 1):
  Col A (1): Country name (English)
  Col B (2): Euro-super 95 (E5)            EUR per 1000 L
  Col C (3): Automotive gas oil (Diesel)   EUR per 1000 L
  Col D (4): Heating gas oil               EUR per 1000 L
  Col E (5): Fuel oil (low sulphur)        EUR per 1000 L
  Col F (6): Fuel oil (high sulphur)       EUR per 1000 L
  Col G (7): LPG motor fuel                EUR per 1000 L

Non-euro countries have additional exchange-rate columns appended; this
provider reads only the first 7 columns (EUR prices) and ignores the rest.

Price normalisation
-------------------
All prices in the Excel file are in EUR per 1000 litres.  This provider
divides every price by 1000.0 to produce EUR/litre, consistent with the
StationData contract.

National averages
-----------------
The EC publishes national weighted averages only — no station-level data.
Each country is exposed as a single pseudo-station whose station_id is the
ISO 3166-1 alpha-2 country code (e.g. 'DE', 'FR').  The special rows
'European Union' and 'Euro area' are mapped to 'EU27' and 'EURO' codes
respectively.

Poll cadence
------------
Data is published every Thursday (Wednesday submission cut-off).
POLL_INTERVAL_SECONDS = 7 * 24 * 3600 = 604800 (one week).

Dependencies
------------
Requires ``openpyxl`` for .xlsx parsing.  openpyxl is not bundled with
Home Assistant core but is available in the HA venv and is a declared
dependency in manifest.json.

Confidence score: 6/10
  - UUID stability is assumed (no official documentation).
  - Weekly not daily; suitable for EU-wide trend sensors.
  - Excel adds openpyxl dependency.
  - No station-level data; useless for finding cheap local stations.
"""

from __future__ import annotations

import asyncio
import functools
import io
import logging
import re
from datetime import UTC, datetime

from aiohttp import ClientResponseError, ClientSession, ClientTimeout

from ..const import API_TIMEOUT
from .base import BaseProvider, ProviderError, StationData

_LOGGER = logging.getLogger(__name__)

# Stable EC document download URL.  Replace the UUID segment to switch between
# the with-taxes / without-taxes / history variants.
_BASE_URL = "https://energy.ec.europa.eu/document/download/{uuid}_en"

# UUID for prices_with_taxes (current week, ~14 KB)
_UUID_WITH_TAXES = "264c2d0f-f161-4ea3-a777-78faae59bea0"
# UUID for prices_without_taxes
_UUID_WITHOUT_TAXES = "78311f92-68f8-4b82-b5cf-1293beeaae77"

_DOWNLOAD_URL = _BASE_URL.format(uuid=_UUID_WITH_TAXES)

_TIMEOUT = ClientTimeout(total=max(API_TIMEOUT, 30))  # Excel download may be slow

_HEADERS: dict[str, str] = {
    "User-Agent": "HomeAssistant/2025.1 aiohttp/3.9.1",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*",
}

# Columns (0-indexed from the first data column in the worksheet).
# openpyxl cell.value is read from row.cells; we use row index in the sheet.
# Column positions in the Excel sheet (1-indexed, openpyxl default):
_COL_COUNTRY = 1  # A: country name
_COL_E5 = 2  # B: Euro-super 95 (E5)
_COL_DIESEL = 3  # C: Automotive gas oil (Diesel)
_COL_HEATING = 4  # D: Heating gas oil
_COL_FUEL_OIL_LS = 5  # E: Fuel oil (low sulphur)
_COL_FUEL_OIL_HS = 6  # F: Fuel oil (high sulphur)
_COL_LPG = 7  # G: LPG motor fuel

# Number of header rows to skip (rows 1-2 are labels/units)
_HEADER_ROWS = 2

# Mapping from lowercase country name (as it appears in the Excel) to ISO code.
# The EC uses English country names.  Keys are normalised to lowercase and
# stripped; add new mappings here if the EC renames a country column.
_COUNTRY_NAME_TO_CODE: dict[str, str] = {
    "austria": "AT",
    "belgium": "BE",
    "bulgaria": "BG",
    "croatia": "HR",
    "cyprus": "CY",
    "czech republic": "CZ",
    "czechia": "CZ",
    "denmark": "DK",
    "estonia": "EE",
    "finland": "FI",
    "france": "FR",
    "germany": "DE",
    "greece": "GR",
    "hungary": "HU",
    "ireland": "IE",
    "italy": "IT",
    "latvia": "LV",
    "lithuania": "LT",
    "luxembourg": "LU",
    "malta": "MT",
    "netherlands": "NL",
    "poland": "PL",
    "portugal": "PT",
    "romania": "RO",
    "slovakia": "SK",
    "slovenia": "SI",
    "spain": "ES",
    "sweden": "SE",
    # Aggregate rows
    "european union": "EU27",
    "eu": "EU27",
    "euro area": "EURO",
    "euro-zone": "EURO",
}

# Poll interval: one week (data published every Thursday)
_POLL_INTERVAL = 7 * 24 * 3600


class EuOilBulletinProvider(BaseProvider):
    """Fetch EU national average fuel prices from the EC Weekly Oil Bulletin.

    This provider downloads the weekly Excel file published by the European
    Commission and parses national-average fuel prices for all 27 EU member
    states plus the EU27 and Euro Area aggregate rows.

    No station-level data is available — one pseudo-station per country is
    exposed, identified by its ISO 3166-1 alpha-2 country code (e.g. 'DE').

    Usage in config flow
    --------------------
    CONFIG_MODE='location' is used here not because we need coordinates but
    because the station_id is the country code (not a numeric ID).  The user
    selects a country from the async_list_stations dropdown.
    """

    COUNTRY = "EU"
    PROVIDER_KEY = "eu_oil_bulletin"
    LABEL = "EC Weekly Oil Bulletin (EU)"
    CONFIG_MODE = "location"
    STATION_LOOKUP_MODE = "location_search"
    POLL_INTERVAL_SECONDS = _POLL_INTERVAL
    REQUIRES_API_KEY = False

    CAPABILITIES: frozenset[str] = frozenset(
        {
            # Fuel prices
            "unleaded",  # Euro-super 95 (E5)
            "diesel",  # Automotive gas oil
            "kerosene",  # Heating gas oil
            "lpg",  # LPG motor fuel
            # Station identity (national average pseudo-station)
            "name",  # Country name
            "county",  # Country name (region context)
            # Timing
            "lastupdated",  # Week date from Excel header
            # Source reference
            "source_station_id",
            # Coordinator sentinels
            "last_successful_fetch",
            "data_fetch_problem",
        }
    )

    STATION_ID_HINT = (
        "Enter the ISO 3166-1 alpha-2 country code (e.g. 'DE' for Germany, "
        "'FR' for France). Use 'EU27' for the EU-wide weighted average or "
        "'EURO' for the Euro Area aggregate."
    )

    def __init__(
        self,
        station_id: str = "EU27",
        latitude: float | None = None,
        longitude: float | None = None,
        radius_km: float | None = None,
    ) -> None:
        """Initialise the provider.

        Args:
            station_id:  ISO country code or aggregate key ('EU27', 'EURO').
                         Stored for use in async_fetch.
            latitude:    Unused; accepted for BaseProvider location-mode compat.
            longitude:   Unused; accepted for BaseProvider location-mode compat.
            radius_km:   Unused; accepted for BaseProvider location-mode compat.
        """
        self._station_id = station_id.upper() if station_id else "EU27"
        self._latitude = latitude
        self._longitude = longitude
        self._radius_km = radius_km
        # Cache: (raw_bytes, fetch_timestamp) — avoids re-downloading if HA
        # calls async_fetch multiple times within the same poll cycle.
        self._cached_workbook_bytes: bytes | None = None
        self._cached_fetch_time: datetime | None = None

    # ── Public interface ──────────────────────────────────────────────────────

    async def async_fetch(
        self,
        session: ClientSession,
        station_id: str,
    ) -> StationData:
        """Download and parse the EC Oil Bulletin Excel; return national data.

        Downloads the weekly Excel file (or uses the in-memory cache if the
        file was already fetched this poll cycle).  Parses Sheet1 to extract
        prices for the requested country code.

        Args:
            session:    aiohttp ClientSession provided by the coordinator.
            station_id: ISO country code or 'EU27' / 'EURO'.

        Returns:
            StationData dict with fuel prices in EUR/litre.

        Raises:
            ProviderError: Country not found in the Excel data, or the
                           Excel file cannot be parsed.
        """
        try:
            import openpyxl  # noqa: PLC0415  (lazy import — optional dependency)
        except ImportError as err:
            raise ProviderError(
                "openpyxl is required for the EU Oil Bulletin provider. "
                "Add 'openpyxl' to the integration's manifest.json requirements."
            ) from err

        wb_bytes = await self._fetch_excel(session)
        country_code = (station_id or self._station_id).upper()

        try:
            wb = await asyncio.get_running_loop().run_in_executor(
                None,
                functools.partial(
                    openpyxl.load_workbook,
                    io.BytesIO(wb_bytes),
                    read_only=True,
                    data_only=True,
                ),
            )
        except Exception as err:
            raise ProviderError(
                f"Failed to parse EC Oil Bulletin Excel file: {err}"
            ) from err

        try:
            sheet = wb.active
            if sheet is None:
                raise ProviderError("EC Oil Bulletin Excel file has no active sheet.")

            # Extract week date: row 2 col A has the reference date string.
            # Row 1 col A is the "in EUR" units header — skip it.
            week_label: str | None = None
            try:
                for row_idx, col_idx in ((2, 1), (1, 2)):
                    header_val = sheet.cell(row=row_idx, column=col_idx).value
                    if header_val:
                        candidate = str(header_val).strip()
                        if candidate and candidate.lower() != "in eur":
                            week_label = candidate
                            break
            except Exception:  # noqa: BLE001
                pass
            if not week_label:
                week_label = datetime.now(tz=UTC).date().isoformat()

            rows_parsed = _parse_sheet(sheet, _HEADER_ROWS)
        finally:
            wb.close()

        # Look up the requested country
        record = rows_parsed.get(country_code)
        if record is None:
            available = sorted(rows_parsed.keys())
            raise ProviderError(
                f"Country code '{country_code}' not found in EC Oil Bulletin data. "
                f"Available codes: {available}"
            )

        return _build_station_data(country_code, record, week_label)

    async def async_fetch_station_name(
        self,
        session: ClientSession,
        station_id: str,
    ) -> str | None:
        """Return a display name for the country code, or None.

        Attempts a quick name lookup from the static mapping table without
        downloading the Excel file.

        Args:
            session:    aiohttp ClientSession (not used; name derived locally).
            station_id: ISO country code.

        Returns:
            Country name string, or None.
        """
        code = (station_id or "").upper()
        # Reverse lookup in the name → code map
        for name, iso in _COUNTRY_NAME_TO_CODE.items():
            if iso == code:
                return name.title()
        if code:
            return f"EU ({code})"
        return None

    async def async_list_stations(
        self,
        session: ClientSession,
        **kwargs: object,
    ) -> list[tuple[str, str]]:
        """Return (country_code, display_label) pairs for all available countries.

        Downloads the Excel file and returns all country rows found in the
        data.  Coordinates kwargs are accepted but ignored — the EC data has
        no spatial dimension.

        Args:
            session: aiohttp ClientSession.
            lat:     Ignored.
            lng:     Ignored.
            radius_km: Ignored.

        Returns:
            Sorted list of (country_code, "Country Name") tuples.
            Empty list on any failure.
        """
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError:
            _LOGGER.warning(
                "openpyxl not available; EU Oil Bulletin station list unavailable."
            )
            return []

        # lat/lng checks: if coordinates are provided but are not None,
        # we still serve all countries (no spatial filtering for national avg).
        lat = kwargs.get("lat")
        lng = kwargs.get("lng")
        if lat is not None and lng is not None:
            _LOGGER.debug(
                "async_list_stations called with lat=%s lng=%s; "
                "EU Oil Bulletin returns all countries (national avg only).",
                lat,
                lng,
            )

        try:
            wb_bytes = await self._fetch_excel(session)
        except Exception as err:  # noqa: BLE001
            _LOGGER.debug(
                "EU Oil Bulletin: failed to fetch Excel for station list: %s", err
            )
            return []

        try:
            wb = await asyncio.get_running_loop().run_in_executor(
                None,
                functools.partial(
                    openpyxl.load_workbook,
                    io.BytesIO(wb_bytes),
                    read_only=True,
                    data_only=True,
                ),
            )
        except Exception as err:  # noqa: BLE001
            _LOGGER.debug(
                "EU Oil Bulletin: failed to parse Excel for station list: %s", err
            )
            return []

        try:
            sheet = wb.active
            if sheet is None:
                return []

            rows_parsed = _parse_sheet(sheet, _HEADER_ROWS)
        finally:
            wb.close()

        result: list[tuple[str, str]] = []
        for code, rec in sorted(rows_parsed.items()):
            country_name = rec.get("country_name") or code
            diesel = rec.get("diesel")
            e5 = rec.get("e5")
            parts = []
            if diesel is not None:
                parts.append(f"Diesel €{diesel:.3f}/L")
            if e5 is not None:
                parts.append(f"E5 €{e5:.3f}/L")
            if parts:
                label = f"{country_name} — {', '.join(parts)}"
            else:
                label = country_name
            result.append((code, label))

        return result

    # ── Internal helpers ──────────────────────────────────────────────────────

    async def _fetch_excel(self, session: ClientSession) -> bytes:
        """Download the EC Oil Bulletin Excel file and return raw bytes.

        Uses a simple in-memory cache keyed by fetch timestamp so that
        multiple calls within the same coordinator poll cycle do not
        re-download the file.  The cache is invalidated after one week.

        Args:
            session: aiohttp ClientSession.

        Returns:
            Raw .xlsx bytes.

        Raises:
            ProviderError: HTTP error or empty response body.
        """
        now = datetime.now(tz=UTC)

        # Cache hit: re-use bytes if downloaded within the last 24 hours
        # (coordinator already enforces the weekly poll; this guard prevents
        # re-downloads if async_fetch is called more than once per cycle).
        if (
            self._cached_workbook_bytes is not None
            and self._cached_fetch_time is not None
            and (now - self._cached_fetch_time).total_seconds() < 86400
        ):
            _LOGGER.debug("EU Oil Bulletin: using cached Excel bytes.")
            return self._cached_workbook_bytes

        _LOGGER.debug("EU Oil Bulletin: downloading Excel from %s", _DOWNLOAD_URL)
        try:
            async with session.get(
                _DOWNLOAD_URL,
                headers=_HEADERS,
                timeout=_TIMEOUT,
                allow_redirects=True,
            ) as response:
                response.raise_for_status()
                content_type = response.headers.get("Content-Type", "")
                data = await response.read()
        except ClientResponseError as err:
            raise ProviderError(
                f"EU Oil Bulletin: HTTP {err.status} when downloading Excel file: {err.message}"
            ) from err
        except Exception as err:
            raise ProviderError(
                f"EU Oil Bulletin: network error downloading Excel file: {err}"
            ) from err

        if not data:
            raise ProviderError("EU Oil Bulletin: downloaded Excel file is empty.")

        _LOGGER.debug(
            "EU Oil Bulletin: downloaded %d bytes (Content-Type: %s)",
            len(data),
            content_type,
        )

        self._cached_workbook_bytes = data
        self._cached_fetch_time = now
        return data


# ── Module-level helpers ──────────────────────────────────────────────────────


def _parse_price_per_litre(raw: object) -> float | None:
    """Convert a raw Excel cell value (EUR per 1000 L) to EUR per litre.

    The EC Oil Bulletin publishes all prices as EUR per 1000 litres.
    Divides by 1000.0 to produce EUR/litre.

    Args:
        raw: Cell value from openpyxl — float, int, str, or None.

    Returns:
        Price in EUR/litre rounded to 4 decimal places, or None on failure.
    """
    if raw is None:
        return None
    try:
        val = float(raw)
    except (ValueError, TypeError):
        return None
    if val <= 0:
        return None
    # Prices are in EUR per 1000 L; divide to get EUR/litre
    return round(val / 1000.0, 4)


def _resolve_country_code(country_name: str) -> str | None:
    """Return the ISO code for an EC country name string, or None.

    Args:
        country_name: Raw string from the Excel country column (may be
                      multiline with language variants separated by newlines).

    Returns:
        ISO 3166-1 alpha-2 code, or None if not recognised.
    """
    # The EC uses multiline cells for EU aggregates: e.g.
    # "CE/EC/EG EUR27_2020 (IV)\nMoyenne pondérée\nWeighted average\n..."
    # Try each line separately to catch these.
    for line in country_name.replace("\r", "\n").split("\n"):
        normalised = line.strip().lower()
        if not normalised:
            continue
        result = _COUNTRY_NAME_TO_CODE.get(normalised)
        if result:
            return result
        # Partial match for composite strings like "EUR27_2020 (IV)".
        # Use word-boundary anchoring to prevent "eu" matching "euro area" etc.
        for key, code in sorted(
            _COUNTRY_NAME_TO_CODE.items(), key=lambda x: len(x[0]), reverse=True
        ):
            if re.search(rf"\b{re.escape(key)}\b", normalised) or re.search(
                rf"\b{re.escape(normalised)}\b", key
            ):
                return code
    return None


def _parse_sheet(sheet: object, header_rows: int) -> dict[str, dict]:
    """Parse an openpyxl worksheet into a dict of country code → price record.

    Iterates rows starting after header_rows, reads columns 1-7, resolves
    the country code, and converts prices from EUR/1000L to EUR/litre.

    Args:
        sheet:       openpyxl Worksheet (or ReadOnlyWorksheet).
        header_rows: Number of header rows to skip (default 2 for EC bulletin).

    Returns:
        Dict mapping country code (e.g. 'DE') to a record dict with keys:
          country_name, e5, diesel, heating_oil, fuel_oil_ls, fuel_oil_hs, lpg
    """
    result: dict[str, dict] = {}
    row_num = 0
    for row in sheet.iter_rows(min_row=header_rows + 1, values_only=True):
        row_num += 1
        if not row:
            continue
        country_raw = row[_COL_COUNTRY - 1] if len(row) >= _COL_COUNTRY else None
        if not country_raw:
            continue
        country_name = str(country_raw).strip()
        if not country_name:
            continue

        code = _resolve_country_code(country_name)
        if code is None:
            # Unrecognised row (blank separators, footer notes, etc.)
            _LOGGER.debug(
                "EU Oil Bulletin: unrecognised country row '%s' — skipped.",
                country_name,
            )
            continue

        def _cell(col_idx: int) -> object:
            """Return cell value for 1-indexed column, or None if out of range."""
            idx = col_idx - 1
            return row[idx] if len(row) > idx else None

        record: dict = {
            "country_name": country_name,
            "e5": _parse_price_per_litre(_cell(_COL_E5)),
            "diesel": _parse_price_per_litre(_cell(_COL_DIESEL)),
            "heating_oil": _parse_price_per_litre(_cell(_COL_HEATING)),
            "fuel_oil_ls": _parse_price_per_litre(_cell(_COL_FUEL_OIL_LS)),
            "fuel_oil_hs": _parse_price_per_litre(_cell(_COL_FUEL_OIL_HS)),
            "lpg": _parse_price_per_litre(_cell(_COL_LPG)),
        }
        result[code] = record

    return result


def _build_station_data(
    country_code: str,
    record: dict,
    week_label: str,
) -> StationData:
    """Assemble a StationData dict from a parsed country price record.

    Args:
        country_code: ISO code used as the pseudo-station identifier.
        record:       Dict from _parse_sheet containing fuel prices.
        week_label:   Reference week string from the Excel header.

    Returns:
        Populated StationData dict.
    """
    country_name: str = record.get("country_name") or country_code

    data: StationData = {
        # Fuel prices (EUR/litre, converted from EUR/1000L)
        "unleaded": record.get("e5"),  # Euro-super 95 (E5)
        "diesel": record.get("diesel"),  # Automotive gas oil
        "kerosene": record.get("heating_oil"),  # Heating gas oil → kerosene key
        "lpg": record.get("lpg"),  # LPG motor fuel
        # Station identity (national average pseudo-station)
        "name": country_name,
        "county": country_name,
        # Timing
        "lastupdated": week_label,
        # Source reference
        "source_station_id": country_code,
    }

    _LOGGER.debug(
        "EU Oil Bulletin parsed for %s: unleaded=%s diesel=%s kerosene=%s lpg=%s week=%s",
        country_code,
        data.get("unleaded"),
        data.get("diesel"),
        data.get("kerosene"),
        data.get("lpg"),
        week_label,
    )

    return data
